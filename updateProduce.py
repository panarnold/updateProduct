#! python
# updateProduce.py - script responsible for change price of the product in the workbook with sales data
# X 2020 Arnold Cytrowski

import openpyxl

wb = openpyxl.load_workbook('produceSalex.xlsx')

sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
'Celery': 1.19,
'Lemon': 1.27}

for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row = row_num, column = 1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row = row_num, column = 2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')